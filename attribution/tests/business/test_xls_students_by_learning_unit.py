##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2016 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################

from django.contrib.auth.models import Group, Permission
from django.urls import reverse
from django.test import TestCase
import openpyxl

from base.tests.factories.tutor import TutorFactory
from base.tests.factories.person import PersonFactory
from attribution.business import xls_students_by_learning_unit

ACCESS_DENIED = 401


class XlsStudentsByLearningUnitTest(TestCase):
    def setUp(self):
        Group.objects.create(name="tutors")
        self.tutor = TutorFactory()
        self.tutor.person.user.user_permissions.add(Permission.objects.get(codename="can_access_attribution"))

        self.url = reverse('produce_xls_students', args=['01234567'])
        self.client.force_login(self.tutor.person.user)

    def create_worksheet(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=xls_students_by_learning_unit.COLUMN_REGISTRATION_ID_NO).value = '12345678'
        worksheet.cell(row=2, column=xls_students_by_learning_unit.COLUMN_REGISTRATION_ID_NO).value = 12345678
        worksheet.cell(row=3, column=xls_students_by_learning_unit.COLUMN_REGISTRATION_ID_NO).value = "zzzzz"
        xls_students_by_learning_unit._columns_registration_id_to_text(worksheet)
        return worksheet

    def test_without_being_logged(self):
        self.client.logout()
        response = self.client.post(self.url)

        self.assertRedirects(response, '/login/?next={}'.format(self.url))

    def test_without_permission(self):
        a_person = PersonFactory()
        self.client.force_login(a_person.user)
        response = self.client.post(self.url)

        self.assertEqual(response.status_code, ACCESS_DENIED)
        self.assertTemplateUsed(response, 'access_denied.html')

    def test_columns_registration_id_to_text(self):
        worksheet = self.create_worksheet()

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.col_idx == xls_students_by_learning_unit.COLUMN_REGISTRATION_ID_NO:
                    self.assertEqual(cell.number_format, xls_students_by_learning_unit.OPENPYXL_STRING_FORMAT)