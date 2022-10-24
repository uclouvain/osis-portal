##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2021 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################

from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Color
from openpyxl.styles import Style
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from attribution.business.student_specific_profile import get_type_peps, is_type_course_or_others
from learning_unit.services.learning_unit import LearningUnitTypeEnum

COLUMN_REGISTRATION_ID_NO = 5
STATUS_COL_WIDTH = 10
NOTE_COL_WIDTH = 10
OPENPYXL_STRING_FORMAT = '@'
BORDER_LEFT = Border(
    left=Side(border_style=BORDER_MEDIUM,
              color=Color('FF000000'),
              ),
)
FIRST_COL_PEPS = 'M'


def get_xls(student_list, acronym, academic_year, learning_unit_type):
    xls = _make_xls_list(student_list, learning_unit_type)
    filename = f"{_('student_list')}_{acronym}_{academic_year}.xlsx"
    response = HttpResponse(xls, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename={filename}"
    return response


def _make_xls_list(student_list, learning_unit_type):
    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = str(_("Students"))
    columns = [
        str(_('Program')),
        str(_('Learning unit')),
        str(_('Email')),
        str(_('Student')),
        str(_('Registration id')),
        str(_('State')),
        str(_('January')),
        str(_('State')),
        str(_('June')),
        str(_('State')),
        str(_('September')),
        str(_('Last score')),
        str(_('Type of specific profile')),
        str(_('Extra time')),
        str(_('Large print')),
        str(_('Specific room of examination')),
        str(_('Other educational facilities for exam')),
        str(_('Other educational facilities for courses and practical exercices')),
        str(_('Guide')),
    ]
    worksheet1.append(col for col in columns)
    for student in student_list:
        line_content = [
            student.get('program'),
            student.get('acronym'),
            student.get('email'),
            student.get('name'),
            student.get('registration_id'),
            student.get('january_status'),
            student.get('january_note'),
            student.get('june_status'),
            student.get('june_note'),
            student.get('september_status'),
            student.get('september_note'),
            student.get('last_note')
        ]

        student_specific_profile = student.get('student_specific_profile')

        if student_specific_profile:

            specific_profile_exam_comment, specific_profile_course_comment = _get_arrangement_comments(
                learning_unit_type,
                student_specific_profile
            )

            line_content.extend([
                get_type_peps(student_specific_profile),
                str(
                    student_specific_profile.arrangement_additional_time_text
                    if student_specific_profile.arrangement_additional_time.value
                    and is_type_course_or_others(learning_unit_type) else '-'
                ),
                str(
                    student_specific_profile.arrangement_appropriate_copy_text
                    if student_specific_profile.arrangement_appropriate_copy.value
                    and is_type_course_or_others(learning_unit_type) else '-'
                ),
                str(_('Yes')) if student_specific_profile.arrangement_specific_locale
                and is_type_course_or_others(learning_unit_type) else '-',
                specific_profile_exam_comment,
                specific_profile_course_comment,
                str(student_specific_profile.guide) if student_specific_profile.guide else '-',
            ])
        else:
            line_content.extend(["-", "-", "-", "-", "-", "-", "-"])
        worksheet1.append(line_content)

    _columns_resizing(worksheet1)
    _columns_registration_id_to_text(worksheet1)
    _set_peps_border(worksheet1, len(student_list) + 1)
    _add_filters_on_headers(worksheet1)
    workbook.create_sheet(str(_("Legend")))
    workbook.worksheets[1].append([str(_("Legend"))])
    workbook.worksheets[1].append([str(_("Exam registration state"))])
    workbook.worksheets[1].append(
        [str(_("P - Examen partiel")), "", str(_("PEPS")), str(_("Program for Students with a Specific Profile"))])
    workbook.worksheets[1].append(
        [str(_("I - Première inscription")), "", str(_("DDI")), str(_("Disability, Disorder or Illness Students"))])
    workbook.worksheets[1].append(
        [str(_("Y - Deuxième inscription")), "", str(_("PMR")), str(_("Person with reduced mobility"))])
    workbook.worksheets[1].append(
        [str(_("J - Report de note de janvier vers septembre")), "", str(_("ESHN")),
         str(_("High Level Promising athlete"))])
    workbook.worksheets[1].append(
        [str(_("R - Report de note de la session précédente")), "", str(_("ES")), str(_("Promising athlete"))])
    workbook.worksheets[1].append(
        [str(_("T - Note résultant d’un test")), "", str(_("Copy PEPS")),
         str(_("A4 single-sided, 1.5 line spacing and Arial 14 font, can be replaced by A3 single-sided"))]
    )
    workbook.worksheets[1].append(
        [str(_("V - Evaluation satisfaisante (la note ne compte pas)")), "", str(_("Additional time")),
         str(_("concerns writings and/or oral preparations"))]
    )
    workbook.worksheets[1].append([str(_("W - Evaluation non satisfaisante (la note ne compte pas)"))])
    workbook.worksheets[1].column_dimensions['A'].width = 50
    workbook.worksheets[1].column_dimensions['C'].width = 6
    workbook.worksheets[1].column_dimensions['D'].width = 50
    return save_virtual_workbook(workbook)


def _get_arrangement_comments(learning_unit_type, student_specific_profile):

    if is_type_course_or_others(learning_unit_type):
        exam_comment = student_specific_profile.arrangement_exam_comment
        specific_profile_exam_comment = ';'.join(
            ([exam_comment] if exam_comment else []) + student_specific_profile.arrangement_exam
        ) or '-'

        course_comment = student_specific_profile.arrangement_course_comment
        specific_profile_course_comment = ';'.join(
            ([course_comment] if course_comment else []) + student_specific_profile.arrangement_course
        ) or '-'

    elif learning_unit_type == LearningUnitTypeEnum.INTERNSHIP.value:
        specific_profile_exam_comment = '-'
        specific_profile_course_comment = student_specific_profile.arrangement_internship_comment or '-'

    elif learning_unit_type == LearningUnitTypeEnum.DISSERTATION.value:
        specific_profile_exam_comment = student_specific_profile.arrangement_dissertation_comment or '-'
        specific_profile_course_comment = '-'

    else:
        specific_profile_exam_comment = '-'
        specific_profile_course_comment = '-'

    return specific_profile_exam_comment, specific_profile_course_comment


def _columns_resizing(ws):
    """
    Definition of the columns sizes
    """
    col_program = ws.column_dimensions['A']
    col_program.width = 18
    col_acronym = ws.column_dimensions['B']
    col_acronym.width = 15
    col_mail = ws.column_dimensions['C']
    col_mail.width = 40
    col_name = ws.column_dimensions['D']
    col_name.width = 30
    col_registration_id = ws.column_dimensions['E']
    col_registration_id.width = 15
    col_january_status = ws.column_dimensions['F']
    col_january_status.width = STATUS_COL_WIDTH
    col_january_note = ws.column_dimensions['G']
    col_january_note.width = NOTE_COL_WIDTH
    col_june_status = ws.column_dimensions['H']
    col_june_status.width = STATUS_COL_WIDTH
    col_june_note = ws.column_dimensions['I']
    col_june_note.width = NOTE_COL_WIDTH
    col_september_status = ws.column_dimensions['J']
    col_september_status.width = STATUS_COL_WIDTH
    col_september_note = ws.column_dimensions['K']
    col_september_note.width = NOTE_COL_WIDTH
    col_last_note = ws.column_dimensions['L']
    col_last_note.width = 15
    col_type_of_specific_profile = ws.column_dimensions['M']
    col_type_of_specific_profile.width = 20
    col_extra_time = ws.column_dimensions['N']
    col_extra_time.width = 25
    col_large_print = ws.column_dimensions['O']
    col_large_print.width = 15
    col_specific_room_of_examination = ws.column_dimensions['P']
    col_specific_room_of_examination.width = 25
    col_educational_tutor = ws.column_dimensions['S']
    col_educational_tutor.width = 30
    col_arrangement_exam = ws.column_dimensions['Q']
    col_arrangement_exam.width = 30
    col_arrangement_course = ws.column_dimensions['R']
    col_arrangement_course.width = 30


def _columns_registration_id_to_text(ws):
    """
    Necessary, otherwise the registration_id is considered as a number and set with a quote while looking at the
    input line
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.col_idx == COLUMN_REGISTRATION_ID_NO:
                cell.number_format = OPENPYXL_STRING_FORMAT


def _set_peps_border(ws, last_row_number):
    """
    Set border at the left of the first peps column
    """
    for cpt in range(1, last_row_number + 1):
        cell = ws[f"{FIRST_COL_PEPS}{cpt}"]
        _update_border_for_first_peps_column(cell)


def _update_border_for_first_peps_column(cell):
    c = cell.style if cell.has_style else Style()
    c.border = BORDER_LEFT
    cell.style = c


def _add_filters_on_headers(a_worksheet):
    full_range = f"A1:{get_column_letter(a_worksheet.max_column)}{str(a_worksheet.max_row)}"
    a_worksheet.auto_filter.ref = full_range
